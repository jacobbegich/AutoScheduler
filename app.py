import streamlit as st
import pandas as pd
import io
from datetime import datetime, timedelta
from pulp import LpProblem, LpVariable, LpMinimize, lpSum, LpBinary, LpStatus
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors

DAYS = ["M", "T", "W", "TH", "F", "SAT", "SUN"]
SHIFTS = ["4am-12pm", "11am-7pm", "6pm-2am"]
EMPLOYEE_COL = "Employee"
STORE_PREFERENCE_COL = "Store Preference"  # Column E - soft preference
HARD_PREFERENCE_COL = "Hard Preference"    # Column F - strict constraint
CANNOT_WORK_WITH_COL = "Cannot Work With"  # Column G - employees who cannot work together

def get_date_range(start_date, end_date):
    """Convert date range to list of dates and day mappings"""
    dates = []
    day_to_dates = {}
    
    current_date = start_date
    while current_date <= end_date:
        # Get day of week (0=Monday, 6=Sunday)
        day_of_week = current_date.weekday()
        
        # Map to our day abbreviations
        day_mapping = {0: "M", 1: "T", 2: "W", 3: "TH", 4: "F", 5: "SAT", 6: "SUN"}
        day_abbrev = day_mapping[day_of_week]
        
        # Store the date
        date_str = current_date.strftime("%Y-%m-%d")
        dates.append(date_str)
        
        # Map day abbreviation to actual dates
        if day_abbrev not in day_to_dates:
            day_to_dates[day_abbrev] = []
        day_to_dates[day_abbrev].append(date_str)
        
        current_date += timedelta(days=1)
    
    return dates, day_to_dates

def convert_availability_to_dates(availability, day_to_dates, days_unavailable, start_date, end_date):
    """Convert availability from day-based to date-based, excluding days_unavailable"""
    date_availability = {}
    # Build a set of unavailable date strings for each employee (YYYY-MM-DD)
    unavailable_map = {}
    for emp, days in days_unavailable.items():
        unavailable_dates = set()
        for d in days:
            try:
                # Parse MM/DD to date in current year
                dt = datetime.strptime(d+f'/{start_date.year}', "%m/%d/%Y")
                # Only include if in range
                if start_date <= dt <= end_date:
                    unavailable_dates.add(dt.strftime("%Y-%m-%d"))
            except Exception:
                continue
        unavailable_map[emp] = unavailable_dates
    for emp, emp_availability in availability.items():
        date_availability[emp] = {}
        for (day, shift, store), available in emp_availability.items():
            if available and day in day_to_dates:
                for date in day_to_dates[day]:
                    if date not in unavailable_map.get(emp, set()):
                        date_availability[emp][(date, shift, store)] = 1
    return date_availability

def parse_availability(df, stores):
    availability = {}
    store_preferences = {}
    hard_preferences = {}
    cannot_work_with = {}
    days_unavailable = {}
    for _, row in df.iterrows():
        emp = row[EMPLOYEE_COL]
        availability[emp] = {}
        preference = row[STORE_PREFERENCE_COL] if STORE_PREFERENCE_COL in df.columns else None
        if pd.notna(preference) and str(preference).strip():
            pref_list = [p.strip() for p in str(preference).split(",") if p.strip()]
            store_preferences[emp] = pref_list if pref_list else None
        else:
            store_preferences[emp] = None
        hard_pref = row[HARD_PREFERENCE_COL] if HARD_PREFERENCE_COL in df.columns else None
        if pd.notna(hard_pref) and str(hard_pref).strip():
            hard_pref_list = [p.strip() for p in str(hard_pref).split(",") if p.strip()]
            hard_preferences[emp] = hard_pref_list if hard_pref_list else None
        else:
            hard_preferences[emp] = None
        cannot_work = row[CANNOT_WORK_WITH_COL] if CANNOT_WORK_WITH_COL in df.columns else None
        if pd.notna(cannot_work) and str(cannot_work).strip():
            cannot_work_list = [p.strip() for p in str(cannot_work).split(",") if p.strip()]
            cannot_work_with[emp] = cannot_work_list if cannot_work_list else None
        else:
            cannot_work_with[emp] = None
        # Parse days unavailable
        days_unavail = row['Days Unavailable (MM/DD)'] if 'Days Unavailable (MM/DD)' in df.columns else None
        if pd.notna(days_unavail) and str(days_unavail).strip():
            days_unavail_list = [d.strip() for d in str(days_unavail).split(",") if d.strip()]
            days_unavailable[emp] = days_unavail_list if days_unavail_list else []
        else:
            days_unavailable[emp] = []
        for shift in SHIFTS:
            days = str(row[shift]).replace(" ","").split(",") if pd.notna(row[shift]) else []
            for day in days:
                if day:
                    for store in stores:
                        availability[emp][(day, shift, store)] = 1
    return availability, store_preferences, hard_preferences, cannot_work_with, days_unavailable

def schedule(availability, store_preferences, hard_preferences, cannot_work_with, employees, days, shifts, stores, store_staffing, max_shifts=5):
    prob = LpProblem("StoreShiftScheduling", LpMinimize)
    x = LpVariable.dicts("assign", [(e, d, s, st) for e in employees for d in days for s in shifts for st in stores], 0, 1, LpBinary)
    understaff = LpVariable.dicts("understaff", [(d, s, st) for d in days for s in shifts for st in stores], 0, None, cat='Integer')
    understaffing_penalty = 1000
    preference_penalty = 1
    understaffing_cost = lpSum([understaff[d, s, st] * understaffing_penalty for d in days for s in shifts for st in stores])
    preference_violations = lpSum([
        x[e, d, s, st] * preference_penalty 
        for e in employees 
        for d in days 
        for s in shifts 
        for st in stores 
        if store_preferences.get(e) and st not in store_preferences[e]
    ])
    prob += understaffing_cost + preference_violations
    for d in days:
        for s in shifts:
            for st in stores:
                required_staff = store_staffing[st]
                prob += lpSum([x[e, d, s, st] for e in employees]) + understaff[d, s, st] >= required_staff
                prob += lpSum([x[e, d, s, st] for e in employees]) <= required_staff
                if required_staff == 1:
                    prob += lpSum([x[e, d, s, st] for e in employees]) == 1
                else:
                    prob += lpSum([x[e, d, s, st] for e in employees]) >= 1
    for e in employees:
        prob += lpSum([x[e, d, s, st] for d in days for s in shifts for st in stores]) <= max_shifts
    for e in employees:
        for d in days:
            prob += lpSum([x[e, d, s, st] for s in shifts for st in stores]) <= 1
    for e in employees:
        for d in days:
            for s in shifts:
                for st in stores:
                    if (d, s, st) not in availability[e]:
                        prob += x[e, d, s, st] == 0
    for e in employees:
        if hard_preferences.get(e):
            preferred_stores = hard_preferences[e]
            for d in days:
                for s in shifts:
                    for st in stores:
                        if st not in preferred_stores:
                            prob += x[e, d, s, st] == 0
    # Add constraints to prevent incompatible employees from working together
    for e1 in employees:
        if cannot_work_with.get(e1):
            incompatible_employees = cannot_work_with[e1]
            for e2 in incompatible_employees:
                if e2 in employees:  # Only add constraint if the incompatible employee exists
                    for d in days:
                        for s in shifts:
                            for st in stores:
                                # If e1 is assigned to this shift/store, e2 cannot be assigned to the same shift/store
                                prob += x[e1, d, s, st] + x[e2, d, s, st] <= 1
    prob.solve()
    return x, understaff, LpStatus[prob.status]

def build_schedule_output(x, understaff, employees, days, shifts, stores, store_staffing, store_preferences, hard_preferences, cannot_work_with):
    schedule = []
    for d in days:
        for s in shifts:
            for st in stores:
                assigned = [e for e in employees if x[e, d, s, st].varValue is not None and abs(x[e, d, s, st].varValue - 1) < 1e-3]
                num_assigned = len(assigned)
                missing = int(understaff[d, s, st].varValue) if understaff[d, s, st].varValue is not None else 0
                required = store_staffing[st]
                staffing_status = f"{num_assigned}/{required}"
                if missing > 0:
                    staffing_status += f" (Understaffed by {missing})"
                
                soft_violations = []
                hard_violations = []
                incompatible_violations = []
                
                # Check for incompatible employee violations
                for emp1 in assigned:
                    if cannot_work_with.get(emp1):
                        for emp2 in assigned:
                            if emp2 in cannot_work_with[emp1]:
                                incompatible_violations.append(f"{emp1} and {emp2} cannot work together")
                
                for emp in assigned:
                    if store_preferences.get(emp) and st not in store_preferences[emp]:
                        pref_str = ", ".join(store_preferences[emp])
                        soft_violations.append(f"{emp} (prefers {pref_str})")
                    if hard_preferences.get(emp) and st not in hard_preferences[emp]:
                        hard_pref_str = ", ".join(hard_preferences[emp])
                        hard_violations.append(f"{emp} (HARD: {hard_pref_str})")
                
                soft_flag = "; ".join(soft_violations) if soft_violations else "None"
                hard_flag = "; ".join(hard_violations) if hard_violations else "None"
                incompatible_flag = "; ".join(incompatible_violations) if incompatible_violations else "None"
                
                schedule.append({
                    "Day": d,
                    "Shift": s,
                    "Store": st,
                    "Employees Assigned": ", ".join(assigned),
                    "Staffing": staffing_status,
                    "Missing Staff": missing,
                    "Soft Preference Violations": soft_flag,
                    "Hard Preference Violations": hard_flag,
                    "Incompatible Employee Violations": incompatible_flag
                })
    return schedule

def build_user_friendly_schedule(x, employees, days, shifts, stores):
    row_tuples = []
    for store in stores:
        for shift in shifts:
            row_tuples.append((store, shift))
    row_index = pd.MultiIndex.from_tuples(row_tuples, names=["Store", "Shift"])
    columns = days
    schedule_df = pd.DataFrame(index=row_index, columns=columns)
    for store in stores:
        for shift in shifts:
            for day in days:
                assigned = [e for e in employees if x[e, day, shift, store].varValue is not None and abs(x[e, day, shift, store].varValue - 1) < 1e-3]
                schedule_df.loc[(store, shift), day] = ", ".join(assigned) if assigned else ""
    return schedule_df

def build_employee_summary(x, employees, days, shifts, stores):
    """Build a summary of total shifts per employee with weekly breakdown"""
    employee_shifts = {}
    employee_weekly_shifts = {}
    
    # Initialize weekly shifts tracking
    for emp in employees:
        employee_weekly_shifts[emp] = {}
    
    for emp in employees:
        total_shifts = 0
        for day in days:
            # Calculate which calendar week this date belongs to
            date_obj = datetime.strptime(day, "%Y-%m-%d")
            # Get the start of the week (Monday) for this date
            week_start = date_obj - timedelta(days=date_obj.weekday())
            week_key = week_start.strftime("%Y-%m-%d")
            
            # Initialize week if not exists
            if week_key not in employee_weekly_shifts[emp]:
                employee_weekly_shifts[emp][week_key] = 0
            
            for shift in shifts:
                for store in stores:
                    if x[emp, day, shift, store].varValue is not None and abs(x[emp, day, shift, store].varValue - 1) < 1e-3:
                        total_shifts += 1
                        employee_weekly_shifts[emp][week_key] += 1
        
        employee_shifts[emp] = total_shifts
    
    # Create the summary DataFrame
    summary_data = {
        "Employee": list(employee_shifts.keys()),
        "Total Shifts": list(employee_shifts.values())
    }
    
    # Add weekly columns
    all_weeks = set()
    for emp_weeks in employee_weekly_shifts.values():
        all_weeks.update(emp_weeks.keys())
    
    # Sort weeks chronologically
    all_weeks = sorted(all_weeks)
    
    for week in all_weeks:
        week_label = f"Week of {datetime.strptime(week, '%Y-%m-%d').strftime('%b %d')}"
        summary_data[week_label] = []
        for emp in employee_shifts.keys():
            summary_data[week_label].append(employee_weekly_shifts[emp].get(week, 0))
    
    summary_df = pd.DataFrame(summary_data)
    return summary_df

def get_availability_template(stores):
    columns = [EMPLOYEE_COL, STORE_PREFERENCE_COL, HARD_PREFERENCE_COL, CANNOT_WORK_WITH_COL, 'Days Unavailable (MM/DD)'] + SHIFTS
    df = pd.DataFrame(columns=columns)
    return df

def run_scheduler(uploaded_file, store_staffing, max_shifts, stores, start_date, end_date):
    df = pd.read_excel(uploaded_file)
    if STORE_PREFERENCE_COL not in df.columns:
        df[STORE_PREFERENCE_COL] = None
    if HARD_PREFERENCE_COL not in df.columns:
        df[HARD_PREFERENCE_COL] = None
    if CANNOT_WORK_WITH_COL not in df.columns:
        df[CANNOT_WORK_WITH_COL] = None
    
    # Parse availability (still day-based)
    availability, store_preferences, hard_preferences, cannot_work_with, days_unavailable = parse_availability(df, stores)
    
    # Convert to date-based availability, excluding days_unavailable
    dates, day_to_dates = get_date_range(start_date, end_date)
    date_availability = convert_availability_to_dates(availability, day_to_dates, days_unavailable, start_date, end_date)
    
    employees = df[EMPLOYEE_COL].tolist()
    
    # Group dates by week
    weeks = {}
    for date in dates:
        date_obj = datetime.strptime(date, "%Y-%m-%d")
        week_start = date_obj - timedelta(days=date_obj.weekday())
        week_key = week_start.strftime("%Y-%m-%d")
        if week_key not in weeks:
            weeks[week_key] = []
        weeks[week_key].append(date)
    
    # Sort weeks chronologically
    sorted_weeks = sorted(weeks.keys())
    
    # Schedule week by week
    all_assignments = {}
    all_understaff = {}
    total_employee_shifts = {emp: 0 for emp in employees}
    
    for week_idx, week_start in enumerate(sorted_weeks):
        week_dates = sorted(weeks[week_start])
        
        # Create week-specific availability
        week_availability = {}
        for emp in employees:
            week_availability[emp] = {}
            for date in week_dates:
                for shift in SHIFTS:
                    for store in stores:
                        if (date, shift, store) in date_availability[emp]:
                            week_availability[emp][(date, shift, store)] = 1
        
        # Adjust max_shifts for this week based on previous weeks
        remaining_weeks = len(sorted_weeks) - week_idx
        weekly_max_shifts = max_shifts  # Keep original weekly limit
        
        # Try to schedule this week
        x, understaff, status = schedule(week_availability, store_preferences, hard_preferences, cannot_work_with, 
                                       employees, week_dates, SHIFTS, stores, store_staffing, weekly_max_shifts)
        
        if status != "Optimal":
            return None, None, None, None, f"No feasible schedule found for week starting {week_start}. Try adjusting preferences or availability."
        
        # Store results for this week
        for emp in employees:
            for date in week_dates:
                for shift in SHIFTS:
                    for store in stores:
                        all_assignments[emp, date, shift, store] = x[emp, date, shift, store]
                        if x[emp, date, shift, store].varValue is not None and abs(x[emp, date, shift, store].varValue - 1) < 1e-3:
                            total_employee_shifts[emp] += 1
        
        for date in week_dates:
            for shift in SHIFTS:
                for store in stores:
                    all_understaff[date, shift, store] = understaff[date, shift, store]
    
    # Build final outputs
    schedule_result = build_schedule_output(all_assignments, all_understaff, employees, dates, SHIFTS, stores, 
                                         store_staffing, store_preferences, hard_preferences, cannot_work_with)
    out_df = pd.DataFrame(schedule_result, columns=["Day", "Shift", "Store", "Employees Assigned", "Staffing", "Missing Staff", "Soft Preference Violations", "Hard Preference Violations", "Incompatible Employee Violations"])
    user_friendly_df = build_user_friendly_schedule(all_assignments, employees, dates, SHIFTS, stores)
    employee_summary_df = build_employee_summary(all_assignments, employees, dates, SHIFTS, stores)
    
    # Generate PDF schedule
    pdf_buffer = generate_pdf_schedule(all_assignments, employees, dates, SHIFTS, stores, start_date, end_date)
    
    return out_df, user_friendly_df, employee_summary_df, pdf_buffer, None

def generate_pdf_schedule(x, employees, dates, shifts, stores, start_date, end_date):
    """Generate a PDF schedule with separate pages for each store in landscape orientation"""
    
    # Create PDF buffer with landscape orientation
    pdf_buffer = io.BytesIO()
    # Use landscape orientation by rotating A4
    landscape_pagesize = A4[1], A4[0]  # Swap width and height for landscape
    doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape_pagesize, rightMargin=0.15*inch, leftMargin=0.15*inch, 
                           topMargin=0.15*inch, bottomMargin=0.15*inch)
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        spaceAfter=15,
        alignment=1  # Center alignment
    )
    
    story = []
    
    # Generate a page for each store
    for store in stores:
        # Store title
        story.append(Paragraph(f"{store} Schedule", title_style))
        story.append(Spacer(1, 0.1*inch))
        
        # Group dates by week
        weeks = {}
        for date in dates:
            date_obj = datetime.strptime(date, "%Y-%m-%d")
            week_start = date_obj - timedelta(days=date_obj.weekday())
            week_key = week_start.strftime("%Y-%m-%d")
            if week_key not in weeks:
                weeks[week_key] = []
            weeks[week_key].append(date)
        
        # Sort weeks chronologically
        sorted_weeks = sorted(weeks.keys())
        
        # Create schedule table for this store
        table_data = []
        
        # Now fill in the data by week
        for week_idx, week in enumerate(sorted_weeks):
            week_dates = sorted(weeks[week])  # Sort dates within week
            
            # Add week header row
            week_start_date = datetime.strptime(week, "%Y-%m-%d")
            week_header = [f"Week of {week_start_date.strftime('%b %d')}"]
            for date in week_dates:
                date_obj = datetime.strptime(date, "%Y-%m-%d")
                week_header.append(date_obj.strftime("%a %b %d"))
            # Pad with empty cells if week has fewer than 7 days
            while len(week_header) < 8:  # 1 for label + 7 days
                week_header.append("")
            table_data.append(week_header)
            
            # Add data rows for this week
            for shift_idx, shift in enumerate(shifts):
                row = [shift]
                for date in week_dates:
                    # Get employees assigned to this store/shift/date
                    assigned = [e for e in employees if x[e, date, shift, store].varValue is not None and 
                               abs(x[e, date, shift, store].varValue - 1) < 1e-3]
                    row.append(", ".join(assigned) if assigned else "")
                # Pad with empty cells if week has fewer than 7 days
                while len(row) < 8:  # 1 for shift + 7 days
                    row.append("")
                table_data.append(row)
            
            # Remove spacing row between weeks
        
        # Calculate column widths for landscape layout
        col_widths = [1.5*inch] + [1.2*inch] * 7  # Shift column wider, 7 day columns
        
        # Create table with landscape-optimized spacing
        table = Table(table_data, colWidths=col_widths)
        
        # Style the table with borders and spacing optimized for landscape
        table_style = TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),  # Smaller font for data cells
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, -1), colors.white),  # All cells white by default
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('WORDWRAP', (0, 0), (-1, -1), True),  # Enable word wrapping
            ('LEADING', (0, 0), (-1, -1), 10),  # Line spacing for wrapped text
        ])
        
        # Style week header rows with light gray
        week_header_rows = []
        for i, week in enumerate(sorted_weeks):
            week_header_row = i * (len(shifts) + 1)  # Calculate week header row position (removed +2 since no spacing row)
            week_header_rows.append(week_header_row)
        
        for row_idx in week_header_rows:
            table_style.add('BACKGROUND', (0, row_idx), (-1, row_idx), colors.lightgrey)
            table_style.add('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold')
            table_style.add('FONTSIZE', (0, row_idx), (-1, row_idx), 8)
            table_style.add('LEFTPADDING', (0, row_idx), (-1, row_idx), 3)
            table_style.add('RIGHTPADDING', (0, row_idx), (-1, row_idx), 3)
            table_style.add('TOPPADDING', (0, row_idx), (-1, row_idx), 2)
            table_style.add('BOTTOMPADDING', (0, row_idx), (-1, row_idx), 2)
        
        # Style shift names in first column with bold
        for week_idx, week in enumerate(sorted_weeks):
            for shift_idx, shift in enumerate(shifts):
                data_row = week_idx * (len(shifts) + 1) + 1 + shift_idx  # Calculate data row position (removed +2 since no spacing row)
                table_style.add('FONTNAME', (0, data_row), (0, data_row), 'Helvetica-Bold')  # Bold first column only
                table_style.add('FONTSIZE', (0, data_row), (-1, data_row), 7)  # Smaller font for data rows
                table_style.add('LEFTPADDING', (0, data_row), (-1, data_row), 3)
                table_style.add('RIGHTPADDING', (0, data_row), (-1, data_row), 3)
                table_style.add('TOPPADDING', (0, data_row), (-1, data_row), 2)
                table_style.add('BOTTOMPADDING', (0, data_row), (-1, data_row), 2)
        
        table.setStyle(table_style)
        
        # Add table to story
        story.append(table)
        
        # Add some space after table
        story.append(Spacer(1, 0.1*inch))
        
        # Remove Notes header and just add blank space for manual notes
        story.append(Spacer(1, 0.5*inch))
        
        # Add page break (except for last store)
        if store != stores[-1]:
            story.append(PageBreak())
    
    # Build PDF
    doc.build(story)
    pdf_buffer.seek(0)
    return pdf_buffer

def main():
    st.title("Store Staff Scheduler")
    st.write("Upload your availability spreadsheet to generate a staff schedule.")
    
    # Store Configuration Section
    st.header("Store Configuration")
    st.write("Configure your stores and staffing requirements:")
    
    # Store input
    store_input = st.text_area(
        "Enter store names (one per line):",
        value="AK&CO\nBookstore\nAK Mercantile\nCabin\nMoosetique",
        help="Enter each store name on a separate line"
    )
    
    # Parse stores from input
    stores = [store.strip() for store in store_input.split('\n') if store.strip()]
    
    if not stores:
        st.error("Please enter at least one store name.")
        return
    
    # Staffing configuration
    st.subheader("Staffing Requirements")
    st.write("Set the number of employees needed per shift at each store:")
    store_staffing = {}
    for store in stores:
        store_staffing[store] = st.number_input(
            f"{store} staff needed per shift", 
            min_value=1, 
            max_value=10, 
            value=2 if store in ["AK&CO", "Bookstore", "AK Mercantile"] else 1,  # Default values
            step=1
        )
    
    # Other parameters
    max_shifts = st.number_input("Maximum shifts per employee per week", min_value=1, max_value=21, value=5, step=1)
    
    # Date range input
    st.subheader("Schedule Date Range")
    st.write("Select the date range for the schedule:")
    st.write("**Note**: The availability template uses day abbreviations (M, T, W, TH, F, SAT, SUN), but the schedule will show actual dates. The system will automatically map your day-based availability to the selected date range.")
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Start Date", value=datetime.now().date())
    with col2:
        end_date = st.date_input("End Date", value=(datetime.now() + timedelta(days=6)).date())
    
    # Show date range info
    if start_date and end_date:
        num_days = (end_date - start_date).days + 1
        st.write(f"**Schedule will cover {num_days} days** ({start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')})")
    
    # Validate date range
    if start_date > end_date:
        st.error("Start date must be before or equal to end date.")
        return
    
    # Convert to datetime objects
    start_datetime = datetime.combine(start_date, datetime.min.time())
    end_datetime = datetime.combine(end_date, datetime.min.time())
    
    # Template download
    st.header("Availability Template")
    st.write("Download a template for your employees to fill out their availability:")
    st.write("**Template columns:**")
    st.write("- **Employee**: Employee name")
    st.write("- **Store Preference**: Preferred stores (comma-separated, soft preference)")
    st.write("- **Hard Preference**: Required stores only (comma-separated, strict constraint)")
    st.write("- **Cannot Work With**: Employees who cannot work together (comma-separated, strict constraint)")
    st.write("- **Days Unavailable (MM/DD)**: Specific dates this employee cannot work (comma-separated, e.g. 06/12, 06/15)")
    st.write("- **Shift columns**: Available days for each shift (comma-separated)")
    st.write("**Note**: Availability uses day abbreviations (M, T, W, TH, F, SAT, SUN) but the schedule will show actual dates. Use 'Days Unavailable' for specific days off within the range.")
    template_df = get_availability_template(stores)
    template_bytes = io.BytesIO()
    template_df.to_excel(template_bytes, index=False)
    template_bytes.seek(0)
    st.download_button(
        label="Download availability template (Excel)",
        data=template_bytes,
        file_name="availability_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    # File upload
    st.header("Generate Schedule")
    uploaded_file = st.file_uploader("Upload availability.xlsx", type=["xlsx"])
    
    if uploaded_file is not None:
        with st.spinner("Generating schedule..."):
            schedule_df, user_friendly_df, employee_summary_df, pdf_buffer, error = run_scheduler(uploaded_file, store_staffing, max_shifts, stores, start_datetime, end_datetime)
        if error:
            st.error(error)
        else:
            st.success(f"Schedule generated for {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}!")
            st.subheader(f"Detailed Schedule ({start_date.strftime('%b %d')} - {end_date.strftime('%b %d')})")
            st.dataframe(schedule_df)
            st.subheader(f"User-Friendly Schedule ({start_date.strftime('%b %d')} - {end_date.strftime('%b %d')})")
            st.dataframe(user_friendly_df)
            st.subheader(f"Employee Summary ({start_date.strftime('%b %d')} - {end_date.strftime('%b %d')})")
            st.dataframe(employee_summary_df)
            towrite = io.BytesIO()
            with pd.ExcelWriter(towrite, engine="openpyxl") as writer:
                schedule_df.to_excel(writer, index=False, sheet_name="Detailed Schedule")
                user_friendly_df.to_excel(writer, sheet_name="User-Friendly Schedule")
                employee_summary_df.to_excel(writer, index=False, sheet_name="Employee Summary")
                
                # Apply conditional formatting to Employee Summary sheet
                workbook = writer.book
                worksheet = workbook["Employee Summary"]
                
                # Create red fill pattern for cells with more than 5 shifts
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                
                # Only apply formatting to weekly columns (columns C onwards)
                for col_idx in range(3, len(employee_summary_df.columns) + 1):  # Start from column C
                    col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                    for row in range(2, len(employee_summary_df) + 2):
                        cell = worksheet[f"{col_letter}{row}"]
                        if cell.value and cell.value >= 6:
                            cell.fill = red_fill
            towrite.seek(0)
            st.download_button(
                label="Download schedule.xlsx",
                data=towrite,
                file_name=f"schedule_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # PDF download
            if pdf_buffer:
                st.download_button(
                    label="Download PDF schedule",
                    data=pdf_buffer,
                    file_name=f"schedule_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf",
                    mime="application/pdf"
                )

if __name__ == "__main__":
    main() 
